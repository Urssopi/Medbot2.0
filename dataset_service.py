import hashlib
import json
import os
import re
from typing import Any

import numpy as np
from openpyxl import load_workbook

try:
    import faiss
except Exception:  # pragma: no cover
    faiss = None  # type: ignore[assignment]

try:
    from openai import OpenAI
except Exception:  # pragma: no cover
    OpenAI = None  # type: ignore[assignment]


class DeidentifiedDataset:
    def __init__(
        self,
        candidate_paths: list[str],
        faiss_index_path: str = "medbot_faiss.index",
        embedding_model: str = "text-embedding-3-small",
        chunk_size_chars: int = 800,
        chunk_overlap_chars: int = 120,
        embedding_batch_size: int = 64,
    ) -> None:
        self.candidate_paths = candidate_paths
        self.faiss_index_path = faiss_index_path
        self.embedding_model = embedding_model
        self.chunk_size_chars = max(200, int(chunk_size_chars))
        self.chunk_overlap_chars = max(0, min(int(chunk_overlap_chars), self.chunk_size_chars - 1))
        self.embedding_batch_size = max(1, int(embedding_batch_size))

        self.records: list[dict[str, Any]] = []
        self.path: str = ""
        self.vector_ready: bool = False
        self.indexed_chunks: int = 0
        self.last_error: str = ""

        self._index: Any = None
        self._chunk_rows: list[dict[str, Any]] = []
        self._meta_path = f"{self.faiss_index_path}.meta.json"
        self._chunks_path = f"{self.faiss_index_path}.chunks.json"

    def _clean_text(self, value: Any) -> str:
        if value is None:
            return ""
        text = str(value)
        return re.sub(r"\s+", " ", text).strip()

    def _truncate(self, text: str, max_len: int) -> str:
        if len(text) <= max_len:
            return text
        return text[: max_len - 3].rstrip() + "..."

    def _select_path(self) -> str:
        for candidate in self.candidate_paths:
            if os.path.exists(candidate):
                return candidate
        return ""

    def _chunk_text(self, text: str) -> list[str]:
        cleaned = self._clean_text(text)
        if not cleaned:
            return []

        step = max(1, self.chunk_size_chars - self.chunk_overlap_chars)
        chunks: list[str] = []
        for start in range(0, len(cleaned), step):
            chunk = cleaned[start : start + self.chunk_size_chars].strip()
            if chunk:
                chunks.append(chunk)
            if start + self.chunk_size_chars >= len(cleaned):
                break
        return chunks

    def _load_cases(self, dataset_path: str) -> list[dict[str, Any]]:
        wb = load_workbook(dataset_path, read_only=True, data_only=True)
        sheet_name = "INCIDENTS" if "INCIDENTS" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        rows = ws.iter_rows(min_row=1, values_only=True)
        headers = [self._clean_text(h) for h in next(rows)]
        col_idx = {name: idx for idx, name in enumerate(headers)}

        def get_value(row: tuple[Any, ...], key: str) -> str:
            idx = col_idx.get(key, -1)
            if idx < 0 or idx >= len(row):
                return ""
            return self._clean_text(row[idx])

        cases: list[dict[str, Any]] = []
        for case_idx, row in enumerate(rows):
            encounter_id = get_value(row, "Encounter ID")
            chief_complaint = get_value(row, "Chief Complaint")
            illness_type = get_value(row, "Type of injury/Illness")
            body_part = get_value(row, "Body Part Involved")
            provisional = get_value(row, "Provisional Diagnosis")
            final_dx = get_value(row, "Final Diagnosis")
            plan = get_value(row, "Initial Plan")
            hpi = get_value(row, "HPI")

            summary = " | ".join(
                item
                for item in [
                    f"Encounter ID: {encounter_id}" if encounter_id else "",
                    f"Chief Complaint: {chief_complaint}" if chief_complaint else "",
                    f"Type: {illness_type}" if illness_type else "",
                    f"Body Part: {body_part}" if body_part else "",
                    f"Provisional Dx: {provisional}" if provisional else "",
                    f"Final Dx: {final_dx}" if final_dx else "",
                    f"Plan: {self._truncate(plan, 240)}" if plan else "",
                    f"HPI: {self._truncate(hpi, 220)}" if hpi else "",
                ]
                if item
            )

            chunk_source = "\n".join(
                item
                for item in [
                    f"Encounter ID: {encounter_id}" if encounter_id else "",
                    f"Chief Complaint: {chief_complaint}" if chief_complaint else "",
                    f"Type of injury/illness: {illness_type}" if illness_type else "",
                    f"Body Part Involved: {body_part}" if body_part else "",
                    f"Provisional Diagnosis: {provisional}" if provisional else "",
                    f"Final Diagnosis: {final_dx}" if final_dx else "",
                    f"Initial Plan: {plan}" if plan else "",
                    f"HPI: {hpi}" if hpi else "",
                ]
                if item
            )

            chunks = self._chunk_text(chunk_source)
            if not chunks and summary:
                chunks = [summary]

            cases.append(
                {
                    "case_idx": case_idx,
                    "encounter_id": encounter_id,
                    "chief_complaint": chief_complaint,
                    "final_dx": final_dx,
                    "summary": summary,
                    "chunks": chunks,
                }
            )

        wb.close()
        return cases

    def _compute_index_fingerprint(self, dataset_path: str, total_chunks: int) -> str:
        stat = os.stat(dataset_path)
        raw = "|".join(
            [
                dataset_path,
                str(stat.st_size),
                str(stat.st_mtime_ns),
                self.embedding_model,
                str(self.chunk_size_chars),
                str(self.chunk_overlap_chars),
                str(total_chunks),
            ]
        )
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()

    def _load_meta(self) -> dict[str, Any]:
        if not os.path.exists(self._meta_path):
            return {}
        try:
            with open(self._meta_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}

    def _save_meta(self, payload: dict[str, Any]) -> None:
        try:
            with open(self._meta_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=True, indent=2)
        except Exception:
            pass

    def _load_chunk_rows(self) -> list[dict[str, Any]]:
        if not os.path.exists(self._chunks_path):
            return []
        try:
            with open(self._chunks_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, list) else []
        except Exception:
            return []

    def _save_chunk_rows(self, rows: list[dict[str, Any]]) -> None:
        with open(self._chunks_path, "w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=True)

    def _embed_texts(self, client: Any, texts: list[str]) -> list[list[float]]:
        if OpenAI is None:
            raise RuntimeError("openai package is not installed.")
        response = client.embeddings.create(model=self.embedding_model, input=texts)
        return [list(item.embedding) for item in response.data]

    def _build_chunk_rows(self, cases: list[dict[str, Any]]) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for case in cases:
            for chunk_idx, chunk_text in enumerate(case["chunks"]):
                rows.append(
                    {
                        "case_idx": case["case_idx"],
                        "chunk_idx": chunk_idx,
                        "encounter_id": case["encounter_id"],
                        "chief_complaint": case["chief_complaint"],
                        "final_dx": case["final_dx"],
                        "summary": case["summary"],
                        "chunk_text": chunk_text,
                    }
                )
        return rows

    def _try_reuse_index(self, fingerprint: str, expected_chunks: int) -> bool:
        if faiss is None:
            self.last_error = "faiss-cpu is not installed. Run: pip install faiss-cpu"
            return False

        meta = self._load_meta()
        if (
            meta.get("fingerprint") != fingerprint
            or int(meta.get("indexed_chunks", 0)) != expected_chunks
            or not os.path.exists(self.faiss_index_path)
            or not os.path.exists(self._chunks_path)
        ):
            return False

        try:
            self._index = faiss.read_index(self.faiss_index_path)
            self._chunk_rows = self._load_chunk_rows()
            if len(self._chunk_rows) != expected_chunks:
                return False
            self.indexed_chunks = len(self._chunk_rows)
            self.vector_ready = True
            return True
        except Exception:
            return False

    def _build_faiss_index(self, rows: list[dict[str, Any]], api_key: str) -> tuple[bool, str]:
        if faiss is None:
            return False, "faiss-cpu is not installed. Run: pip install faiss-cpu"
        if not api_key:
            return False, "Missing OPENAI_API_KEY; cannot build query embeddings for RAG."
        if not rows:
            return False, "No chunkable text found in dataset."

        openai_client = OpenAI(api_key=api_key)

        vectors: list[list[float]] = []
        for start in range(0, len(rows), self.embedding_batch_size):
            batch = rows[start : start + self.embedding_batch_size]
            batch_vectors = self._embed_texts(openai_client, [row["chunk_text"] for row in batch])
            vectors.extend(batch_vectors)

        matrix = np.asarray(vectors, dtype="float32")
        if matrix.ndim != 2 or matrix.shape[0] != len(rows):
            return False, "Embedding matrix shape mismatch."

        faiss.normalize_L2(matrix)
        index = faiss.IndexFlatIP(matrix.shape[1])
        index.add(matrix)

        faiss.write_index(index, self.faiss_index_path)
        self._save_chunk_rows(rows)

        self._index = index
        self._chunk_rows = rows
        self.indexed_chunks = len(rows)
        self.vector_ready = True
        return True, f"Indexed {len(rows)} chunks in FAISS."

    def load(self, api_key: str = "", rebuild_index: bool = False) -> tuple[bool, str]:
        self.vector_ready = False
        self.indexed_chunks = 0
        self.last_error = ""

        dataset_path = self._select_path()
        if not dataset_path:
            self.records = []
            self.path = ""
            return False, "Dataset not found."

        self.path = dataset_path
        cases = self._load_cases(dataset_path)
        self.records = [
            {
                "encounter_id": case["encounter_id"],
                "chief_complaint": case["chief_complaint"],
                "final_dx": case["final_dx"],
                "summary": case["summary"],
            }
            for case in cases
        ]

        rows = self._build_chunk_rows(cases)
        total_chunks = len(rows)
        fingerprint = self._compute_index_fingerprint(dataset_path, total_chunks)

        if not rebuild_index and self._try_reuse_index(fingerprint, total_chunks):
            return True, f"Loaded {len(cases)} cases. Reused FAISS index with {total_chunks} chunks."

        try:
            ok, message = self._build_faiss_index(rows, api_key.strip())
            if ok:
                self._save_meta(
                    {
                        "fingerprint": fingerprint,
                        "indexed_chunks": total_chunks,
                        "embedding_model": self.embedding_model,
                        "chunk_size_chars": self.chunk_size_chars,
                        "chunk_overlap_chars": self.chunk_overlap_chars,
                        "dataset_path": self.path,
                    }
                )
                return True, f"Loaded {len(cases)} cases. {message}"
            return False, f"Loaded {len(cases)} cases, but RAG index failed: {message}"
        except Exception as exc:
            self.last_error = str(exc)
            return False, f"Loaded {len(cases)} cases, but RAG index failed: {exc}"

    def search(self, query: str, top_k: int = 5) -> list[dict[str, Any]]:
        query_text = self._clean_text(query)
        if not query_text or top_k <= 0 or not self.vector_ready or self._index is None:
            return []

        api_key = os.getenv("OPENAI_API_KEY", "").strip()
        if not api_key:
            return []

        try:
            openai_client = OpenAI(api_key=api_key)
            qv = np.asarray(self._embed_texts(openai_client, [query_text]), dtype="float32")
            faiss.normalize_L2(qv)

            hit_limit = min(max(top_k * 6, top_k), len(self._chunk_rows))
            scores, ids = self._index.search(qv, hit_limit)

            by_case: dict[str, dict[str, Any]] = {}
            for score, idx in zip(scores[0], ids[0]):
                if idx < 0 or idx >= len(self._chunk_rows):
                    continue
                row = self._chunk_rows[int(idx)]
                case_key = f"{row.get('case_idx', '')}:{row.get('encounter_id', '')}"

                candidate = {
                    "encounter_id": row.get("encounter_id", ""),
                    "chief_complaint": row.get("chief_complaint", ""),
                    "final_dx": row.get("final_dx", ""),
                    "summary": row.get("summary", ""),
                    "chunk_text": row.get("chunk_text", ""),
                    "chunk_excerpt": self._truncate(row.get("chunk_text", ""), 260),
                    "score": round(float(score), 4),
                }

                if case_key not in by_case or candidate["score"] > by_case[case_key]["score"]:
                    by_case[case_key] = candidate

            ranked = sorted(by_case.values(), key=lambda item: item["score"], reverse=True)
            return ranked[:top_k]
        except Exception:
            return []

    def build_context(self, matches: list[dict[str, Any]]) -> str:
        if not matches:
            return "No close matching reference cases found in the local de-identified dataset."

        lines: list[str] = []
        for idx, row in enumerate(matches, start=1):
            summary = row.get("summary", "")
            chunk_excerpt = row.get("chunk_excerpt", "")
            if chunk_excerpt:
                lines.append(f"{idx}. {summary} | Supporting chunk: {chunk_excerpt}")
            else:
                lines.append(f"{idx}. {summary}")
        return "\n".join(lines)
